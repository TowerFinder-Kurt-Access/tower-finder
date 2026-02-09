"""
Fetch parcel owner data from ReportAll API for Alberta towers
Processes the first 300 rows from Alberta_Jan11.xlsx
"""
import openpyxl
import requests
import json
import os
from datetime import datetime
from typing import Dict, Any, List, Optional

# Load environment variables from .env.development.local
def load_env_file(filepath: str = '.env.development.local'):
    """Load environment variables from file"""
    env_vars = {}
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    env_vars[key.strip()] = value.strip()
    return env_vars

# Load API key
env_vars = load_env_file()
API_KEY = env_vars.get('REPORTALL_API_KEY')

if not API_KEY:
    raise ValueError("REPORTALL_API_KEY not found in .env.development.local")

print(f"API Key loaded: {API_KEY[:8]}...")

def fetch_parcel_data(lat: float, lon: float) -> Optional[Dict[str, Any]]:
    """
    Fetch parcel data from ReportAll API for given coordinates
    Similar to InformationService.fetchExternalParcelData()
    """
    try:
        point_wkt = f"POINT({lon} {lat})"
        params = {
            'client': API_KEY,
            'v': 9,
            'spatial_nearest': point_wkt,
            'sn_srid': 4326,
            'limit': 1
        }

        print(f"\n{'='*60}")
        print(f"REPORTALL API REQUEST - {datetime.now().isoformat()}")
        print(f"Coordinates: ({lat}, {lon})")
        print(f"Point WKT: {point_wkt}")
        print(f"{'='*60}")

        response = requests.get(
            'https://reportallusa.com/api/parcels',
            params=params,
            timeout=60
        )

        print(f"Status: {response.status_code}")

        if response.status_code == 200:
            data = response.json()
            results_count = len(data.get('results', []))
            print(f"Results Count: {results_count}")

            if results_count > 0:
                result = data['results'][0]
                print(f"Parcel ID: {result.get('parcel_id', 'N/A')}")
                print(f"Owner: {result.get('owner', 'N/A')}")
                print(f"Address: {result.get('address', 'N/A')}")
                return {
                    'lat': lat,
                    'lon': lon,
                    'success': True,
                    'data': result,
                    'timestamp': datetime.now().isoformat()
                }
            else:
                print("No results found")
                return {
                    'lat': lat,
                    'lon': lon,
                    'success': False,
                    'error': 'No results found',
                    'timestamp': datetime.now().isoformat()
                }
        else:
            print(f"Error: HTTP {response.status_code}")
            return {
                'lat': lat,
                'lon': lon,
                'success': False,
                'error': f"HTTP {response.status_code}",
                'timestamp': datetime.now().isoformat()
            }

    except Exception as e:
        print(f"Exception: {str(e)}")
        return {
            'lat': lat,
            'lon': lon,
            'success': False,
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }

def main():
    """Main execution function"""
    print("\n" + "="*60)
    print("Alberta Parcel Owner Fetcher")
    print("="*60)
    print(f"Start Time: {datetime.now().isoformat()}")
    print(f"Processing: First 300 rows from Alberta_Jan11.xlsx")
    print("="*60 + "\n")

    # Load Excel file
    print("Loading Excel file...")
    workbook = openpyxl.load_workbook('sources/2. Alberta_Jan11.xlsx', read_only=True)
    sheet = workbook.active

    # Get headers
    headers = [cell.value for cell in sheet[1]]
    lat_idx = headers.index('latitude') + 1
    lon_idx = headers.index('longitude') + 1

    print(f"Total rows in file: {sheet.max_row}")
    print(f"Processing rows: 2 to 301 (first 300 data rows)")
    print(f"Latitude column: {lat_idx}, Longitude column: {lon_idx}\n")

    # Collect results
    results: List[Dict[str, Any]] = []

    # Process first 300 rows (rows 2-301, since row 1 is headers)
    for row_num in range(2, min(302, sheet.max_row + 1)):
        lat = sheet.cell(row_num, lat_idx).value
        lon = sheet.cell(row_num, lon_idx).value

        if lat is None or lon is None:
            print(f"\nRow {row_num}: Skipping - missing coordinates")
            results.append({
                'row': row_num,
                'lat': lat,
                'lon': lon,
                'success': False,
                'error': 'Missing coordinates',
                'timestamp': datetime.now().isoformat()
            })
            continue

        print(f"\nProcessing row {row_num}/{301}...")
        result = fetch_parcel_data(float(lat), float(lon))
        result['row'] = row_num
        results.append(result)

        # Progress indicator
        if row_num % 10 == 0:
            success_count = sum(1 for r in results if r.get('success', False))
            print(f"\n>>> Progress: {len(results)}/300 rows processed, {success_count} successful <<<\n")

    workbook.close()

    # Summary statistics
    print("\n" + "="*60)
    print("PROCESSING COMPLETE")
    print("="*60)
    total_count = len(results)
    success_count = sum(1 for r in results if r.get('success', False))
    fail_count = total_count - success_count

    print(f"Total Rows Processed: {total_count}")
    print(f"Successful API Calls: {success_count}")
    print(f"Failed API Calls: {fail_count}")
    print(f"Success Rate: {(success_count/total_count*100):.1f}%")

    # Save results to JSON file
    output_file = 'experiments/alberta_reportall_results.json'
    print(f"\nSaving results to: {output_file}")

    output_data = {
        'metadata': {
            'source_file': 'sources/2. Alberta_Jan11.xlsx',
            'rows_processed': total_count,
            'rows_range': '2-301',
            'successful_calls': success_count,
            'failed_calls': fail_count,
            'success_rate': f"{(success_count/total_count*100):.1f}%",
            'start_time': datetime.now().isoformat(),
            'api_endpoint': 'https://reportallusa.com/api/parcels',
            'api_version': 9
        },
        'results': results
    }

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)

    print(f"✓ Results saved successfully!")
    print(f"End Time: {datetime.now().isoformat()}")
    print("="*60 + "\n")

if __name__ == '__main__':
    main()
