"""Test script - fetch just 5 rows to verify it works"""
import openpyxl
import requests
import json
import os
import sys
from datetime import datetime

# Load API key
def load_env_file(filepath: str = '.env.development.local'):
    env_vars = {}
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    env_vars[key.strip()] = value.strip()
    return env_vars

env_vars = load_env_file()
API_KEY = env_vars.get('REPORTALL_API_KEY')

if not API_KEY:
    print("ERROR: REPORTALL_API_KEY not found in .env.development.local")
    sys.exit(1)

print(f"API Key loaded: {API_KEY[:8]}...")
print("Loading Excel file...")

try:
    workbook = openpyxl.load_workbook('sources/2. Alberta_Jan11.xlsx', read_only=True)
    sheet = workbook.active
    print(f"Excel file loaded successfully! Total rows: {sheet.max_row}")

    # Get headers
    headers = [cell.value for cell in sheet[1]]
    lat_idx = headers.index('latitude') + 1
    lon_idx = headers.index('longitude') + 1

    print(f"\nTesting with first 5 rows...\n")

    results = []
    for row_num in range(2, 7):  # Just 5 rows for testing
        lat = sheet.cell(row_num, lat_idx).value
        lon = sheet.cell(row_num, lon_idx).value

        print(f"Row {row_num}: lat={lat}, lon={lon}")

        if lat and lon:
            point_wkt = f"POINT({lon} {lat})"
            params = {
                'client': API_KEY,
                'v': 9,
                'spatial_nearest': point_wkt,
                'sn_srid': 4326,
                'limit': 1
            }

            try:
                print(f"  Calling API...")
                response = requests.get(
                    'https://reportallusa.com/api/parcels',
                    params=params,
                    timeout=30
                )
                print(f"  Status: {response.status_code}")

                if response.status_code == 200:
                    data = response.json()
                    if data.get('results'):
                        result = data['results'][0]
                        print(f"  ✓ Owner: {result.get('owner', 'N/A')}")
                        print(f"  ✓ Parcel: {result.get('parcel_id', 'N/A')}")
                        results.append({'row': row_num, 'success': True, 'data': result})
                    else:
                        print(f"  ✗ No results found")
                        results.append({'row': row_num, 'success': False, 'error': 'No results'})
                else:
                    print(f"  ✗ HTTP {response.status_code}")
                    results.append({'row': row_num, 'success': False, 'error': f"HTTP {response.status_code}"})

            except Exception as e:
                print(f"  ✗ Error: {str(e)}")
                results.append({'row': row_num, 'success': False, 'error': str(e)})

        print()

    workbook.close()

    print("\n" + "="*60)
    print(f"Test complete! {sum(1 for r in results if r.get('success'))} / 5 successful")
    print("="*60)

    # Save test results
    with open('experiments/test_results.json', 'w') as f:
        json.dump(results, f, indent=2)
    print("Test results saved to experiments/test_results.json")

except Exception as e:
    print(f"ERROR: {str(e)}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
