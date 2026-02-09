import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('sources/2. Alberta_Jan11.xlsx')
sheet = workbook.active

# Print column headers
print("Column Headers:")
headers = []
for cell in sheet[1]:
    headers.append(cell.value)
    print(f"  {cell.column_letter}: {cell.value}")

# Print first 3 rows of data
print("\nFirst 3 data rows:")
for row_idx in range(2, min(5, sheet.max_row + 1)):
    row_data = {}
    for col_idx, header in enumerate(headers, start=1):
        cell_value = sheet.cell(row_idx, col_idx).value
        row_data[header] = cell_value
    print(f"\nRow {row_idx}:")
    for key, value in row_data.items():
        print(f"  {key}: {value}")

print(f"\nTotal rows: {sheet.max_row}")
